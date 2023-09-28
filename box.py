import openpyxl
import warnings
import datetime
import os

def box ():
    warnings.simplefilter("ignore")

    book_new = openpyxl.open("ТК/Отчет НП Боксберри РУ от 03.07.23.xlsx")
    sheet_new = book_new.active

    list_np = []
    sm = 0

    book = openpyxl.open("пул/b1.xlsx")
    sheet = book.active

    for row in range(13, sheet.max_row-2):
        if sheet[row][30].value != None:
            a = int(sheet[row][14].value)
            b = sheet[row][30].value
            sm += b
            if sheet[row][33].value != "Наличные":
                c = 2
            else:
                c = ""
            list_np.append([a, b, c])
    book.close()

    if os.path.exists("пул/b2.xlsx"):
        book = openpyxl.open("пул/b2.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b3.xlsx"):
        book = openpyxl.open("пул/b3.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b4.xlsx"):
        book = openpyxl.open("пул/b4.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b5.xlsx"):
        book = openpyxl.open("пул/b5.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b6.xlsx"):
        book = openpyxl.open("пул/b6.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b7.xlsx"):
        book = openpyxl.open("пул/b7.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b8.xlsx"):
        book = openpyxl.open("пул/b8.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/b9.xlsx"):
        book = openpyxl.open("пул/b9.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    #запись
    for i in range(1, len(list_np) + 1):
        # i принимает значения от 1 до 17, пишем со второй строки I+1, начиная с 0 индекса списка, 0 индекса подсписка I-1
        # 1,len(lst)+1 позваляет перебирать значения от 1 а не от 1, и включительно по число длины списка

        sheet_new[i + 1][0].value = list_np[i - 1][0]
        sheet_new[i + 1][7].value = list_np[i - 1][1]
        sheet_new[i + 1][8].value = list_np[i - 1][2]

    today = datetime.date.today()
    now = datetime.datetime.now()
    weekday = now.weekday()
    if weekday != 0:
        yesterday = today - datetime.timedelta(days=1)
    else:
        yesterday = today - datetime.timedelta(days=3)

    book_new.save("res/Отчет НП Боксберри РУ от %s %s.xlsx" % (yesterday, round(sm, 2)))
    book_new.close()
