import openpyxl
import warnings
import datetime
import os


def log_sr():
    warnings.simplefilter("ignore")

    book = openpyxl.open("пул/o1.xlsx")
    sheet = book.active

    if os.path.exists("пул/o2.xlsx"):
        book22 = openpyxl.open("пул/o2.xlsx")
        sheet22 = book22.active

    if os.path.exists("пул/o3.xlsx"):
        book222 = openpyxl.open("пул/o3.xlsx")
        sheet222 = book222.active

    book_new = openpyxl.open("ТК/Отчет ОРИОН НП от 03.07.23.xlsx")
    sheet_new = book_new.active

    book_new1 = openpyxl.open("ТК/Отчет ОРИОН СЧЕТ 03.07.23.xlsx")
    sheet_new1 = book_new1.active

    sm = 0
    lst_np = []
    lst_sch = []

    ksm1 = 0
    ksm2 = 0
    ksm3 = 0

    # код счета

    for row in range(5, sheet.max_row):
        if sheet[row][2].value != None:
            if sheet[row][6].value != "Отмена":
                a = int(sheet[row][2].value)
                b = sheet[row][1].value
                if sheet[row][11].value != None:
                    c = sheet[row][11].value
                else:
                    c = ""
                d = sheet[row][16].value
                e = [a, b, c, d]
                lst_sch.append(e)
    # print(lst_sch)

    if os.path.exists("пул/o2.xlsx"):
        for row in range(5, sheet22.max_row):
            if sheet22[row][2].value != None:
                if sheet22[row][6].value != "Отмена":
                    a = int(sheet22[row][2].value)
                    b = sheet22[row][1].value
                    if sheet22[row][11].value != None:
                        c = sheet22[row][11].value
                    else:
                        c = ""
                    d = sheet22[row][16].value
                    e = [a, b, c, d]
                    lst_sch.append(e)

    if os.path.exists("пул/o3.xlsx"):
        for row in range(5, sheet222.max_row):
            if sheet222[row][2].value != None:
                if sheet222[row][6].value != "Отмена":
                    a = int(sheet222[row][2].value)
                    b = sheet222[row][1].value
                    if sheet222[row][11].value != None:
                        c = sheet222[row][11].value
                    else:
                        c = ""
                    d = sheet222[row][16].value
                    e = [a, b, c, d]
                    lst_sch.append(e)

    for i in range(1, len(lst_sch) + 1):
        # i принимает значения от 1 до 17, пишем со второй строки I+1, начиная с 0 индекса списка, 0 индекса подсписка I-1
        # 1,len(lst)+1 позваляет перебирать значения от 1 а не от 0, и включительно по число длины списка
        sheet_new1[i + 1][0].value = lst_sch[i - 1][0]
        sheet_new1[i + 1][1].value = lst_sch[i - 1][1]
        sheet_new1[i + 1][17].value = lst_sch[i - 1][2]
        sheet_new1[i + 1][22].value = lst_sch[i - 1][3]

    # код наложки

    for row in range(5, sheet.max_row):
        if sheet[row][2].value != None:
            if sheet[row][6].value != "Отмена":
                if sheet[row][15].value != 0.00:
                    a = int(sheet[row][2].value)
                    b = sheet[row][1].value
                    if sheet[row][7].value != "Нал.":
                        c = 2
                    else:
                        c = ""
                    d = sheet[row][15].value
                    sm += d
                    ksm1 += d
                    e = [a, b, d, c]
                    lst_np.append(e)
    ksm1 = (round(ksm1, 2))

    if os.path.exists("пул/o2.xlsx"):
        for row in range(5, sheet22.max_row):
            if sheet22[row][2].value != None:
                if sheet22[row][6].value != "Отмена":
                    if sheet22[row][15].value != 0.00:
                        a = int(sheet22[row][2].value)
                        b = sheet22[row][1].value
                        if sheet22[row][7].value != "Нал.":
                            c = 2
                        else:
                            c = ""
                        d = sheet22[row][15].value
                        sm += d
                        ksm2 += d
                        e = [a, b, d, c]
                        lst_np.append(e)
        ksm2 = (round(ksm2, 2))

    if os.path.exists("пул/o3.xlsx"):
        for row in range(5, sheet222.max_row):
            if sheet222[row][2].value != None:
                if sheet222[row][6].value != "Отмена":
                    if sheet222[row][15].value != 0.00:
                        a = int(sheet222[row][2].value)
                        b = sheet222[row][1].value
                        if sheet222[row][7].value != "Нал.":
                            c = 2
                        else:
                            c = ""
                        d = sheet222[row][15].value
                        sm += d
                        ksm3 += d
                        e = [a, b, d, c]
                        lst_np.append(e)

        ksm3 = (round(ksm3, 2))

    for i in range(1, len(lst_np) + 1):
        # i принимает значения от 1 до 17, пишем со второй строки I+1, начиная с 0 индекса списка, 0 индекса подсписка I-1
        # 1,len(lst)+1 позваляет перебирать значения от 1 а не от 0, и включительно по число длины списка
        sheet_new[i + 1][0].value = lst_np[i - 1][0]
        sheet_new[i + 1][1].value = lst_np[i - 1][1]
        sheet_new[i + 1][7].value = lst_np[i - 1][2]
        sheet_new[i + 1][8].value = lst_np[i - 1][3]

    today = datetime.date.today()
    now = datetime.datetime.now()
    weekday = now.weekday()
    if weekday != 0:
        yesterday = today - datetime.timedelta(days=1)
    else:
        yesterday = today - datetime.timedelta(days=3)

    book_new1.save("res/Отчет ОРИОН СЧЕТ от %s.xlsx" % (yesterday))
    book_new.save("res/Отчет ОРИОН НП от %s %s.xlsx" % (yesterday, round(sm, 2)))

    book_new1.close()
    book_new.close()
    book.close()

    if os.path.exists("пул/o2.xlsx"):
        book22.close()

    if os.path.exists("пул/o3.xlsx"):
        book222.close()

    ksm = "Контрольные суммы Орион: *%s*  *%s*  *%s*" % (ksm1, ksm2, ksm3)

    return ksm