import openpyxl
import warnings
import datetime
import os

def dpd_safe ():
    warnings.simplefilter("ignore")

    book_new = openpyxl.open("ТК/Отчет ДПД НП СЭЙФ от 03.07.23.xlsx")
    sheet_new = book_new.active

    list_np = []
    sm = 0

    book = openpyxl.open("пул/ds1.xlsx")
    sheet = book.active

    for row in range(8, sheet.max_row-10):
        a = int(sheet[row][3].value)
        if "RU" in sheet[row][2].value:
            b = sheet[row][2].value
        else:
            b = int(sheet[row][2].value)
        d = float(sheet[row][8].value)
        sm += d
        if sheet[row][7].value != "Cash":
            c = 2
        else:
            c = ""
        list_np.append([a, b, d, c])
    book.close()

    if os.path.exists("пул/ds2.xlsx"):
        book = openpyxl.open("пул/ds2.xlsx")
        sheet = book.active

        for row in range(8, sheet.max_row-10):
            a = int(sheet[row][3].value)
            if "RU" in sheet[row][2].value:
                b = sheet[row][2].value
            else:
                b = int(sheet[row][2].value)
            d = float(sheet[row][8].value)
            sm += d
            if sheet[row][7].value != "Cash":
                c = 2
            else:
                c = ""
            list_np.append([a, b, d, c])
        book.close()

    if os.path.exists("пул/ds3.xlsx"):
        book = openpyxl.open("пул/ds3.xlsx")
        sheet = book.active

        for row in range(8, sheet.max_row-10):
            a = int(sheet[row][3].value)
            if "RU" in sheet[row][2].value:
                b = sheet[row][2].value
            else:
                b = int(sheet[row][2].value)
            d = float(sheet[row][8].value)
            sm += d
            if sheet[row][7].value != "Cash":
                c = 2
            else:
                c = ""
            list_np.append([a, b, d, c])
        book.close()

    # запись
    for i in range(1, len(list_np) + 1):
        # i принимает значения от 1 до 17, пишем со второй строки I+1, начиная с 0 индекса списка, 0 индекса подсписка I-1
        # 1,len(lst)+1 позваляет перебирать значения от 1 а не от 1, и включительно по число длины списка

        sheet_new[i + 1][0].value = list_np[i - 1][0]
        sheet_new[i + 1][1].value = list_np[i - 1][1]
        sheet_new[i + 1][7].value = list_np[i - 1][2]
        sheet_new[i + 1][8].value = list_np[i - 1][3]

    today = datetime.date.today()
    now = datetime.datetime.now()
    weekday = now.weekday()
    if weekday != 0:
        yesterday = today - datetime.timedelta(days=1)
    else:
        yesterday = today - datetime.timedelta(days=3)

    book_new.save("res/Отчет ДПД НП СЭЙФ от %s %s.xlsx" % (yesterday, round(sm, 2)))
    book_new.close()