import openpyxl
import warnings
import datetime

def f_pst_sr1 ():

    warnings.simplefilter("ignore")

    book = openpyxl.open("пул/f1.xlsx")
    sheet = book.active

    book_new = openpyxl.open("ТК/Отчет 5ПОСТ НП от 03.07.23.xlsx")
    sheet_new = book_new.active

    sm = 0

    for row in range(2,sheet.max_row):
        sheet_new[row][0].value = int(sheet[row][0].value)
        sheet_new[row][7].value = sheet[row][5].value
        sm += sheet[row][5].value
        if sheet[row][10].value == "Наличный расчет":
            sheet_new[row][8].value = ""
        else:
            sheet_new[row][8].value = 2

    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)

    book_new.save("res/Отчет 5ПОСТ НП от %s %s.xlsx"%(yesterday, round(sm,2)))
    book_new.close()
    book.close()



def f_pst_sr2 ():

    warnings.simplefilter("ignore")

    book = openpyxl.open("пул/f1.xlsx")
    sheet = book.active

    book1 = openpyxl.open("пул/f2.xlsx")
    sheet1 = book1.active

    book_new = openpyxl.open("ТК/Отчет 5ПОСТ НП от 03.07.23.xlsx")
    sheet_new = book_new.active

    sm = 0

    lst = []

    for row in range(2, sheet.max_row):
        a = int(sheet[row][0].value)
        b = sheet[row][5].value
        sm += sheet[row][5].value
        if sheet[row][10].value == "Наличный расчет":
            c = ""
        else:
            c = 2
        d = [a, b, c]
        lst.append(d)

    for row in range(2, sheet1.max_row):
        a = int(sheet1[row][0].value)
        b = sheet1[row][9].value
        sm += sheet1[row][9].value
        if sheet1[row][14].value == "Наличный расчет":
            c = ""
        else:
            c = 2
        d = [a, b, c]
        lst.append(d)

    for i in range(1, len(lst) + 1):
        # i принимает значения от 1 до 17, пишем со второй строки I+1, начиная с 0 индекса списка, 0 индекса подсписка I-1
        # 1,len(lst)+1 позваляет перебирать значения от 1 а не от 1, и включительно по число длины списка

        sheet_new[i + 1][0].value = lst[i - 1][0]
        sheet_new[i + 1][7].value = lst[i - 1][1]
        sheet_new[i + 1][8].value = lst[i - 1][2]

    today = datetime.date.today()
    now = datetime.datetime.now()
    weekday = now.weekday()
    if weekday != 0:
        yesterday = today - datetime.timedelta(days=1)
    else:
        yesterday = today - datetime.timedelta(days=3)

    book_new.save("res/Отчет 5ПОСТ НП от %s %s.xlsx" % (yesterday, round(sm,2)))
    book_new.close()
    book1.close()
    book.close()

