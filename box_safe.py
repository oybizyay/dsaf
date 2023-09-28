import openpyxl
import warnings
import datetime
import os

def box_safe ():
    warnings.simplefilter("ignore")

    book_new = openpyxl.open("ТК/Отчет НП Боксберри РУ СЭЙФ от 03.07.23.xlsx")
    sheet_new = book_new.active

    list_np = []
    sm = 0

    ksm1 = 0
    ksm2 = 0
    ksm3 = 0
    ksm4 = 0
    ksm5 = 0
    ksm6 = 0
    ksm7 = 0
    ksm8 = 0
    ksm9 = 0
    ksm = ["Контрольные суммы Боксберри: "]

    book = openpyxl.open("пул/bs1.xlsx")
    sheet = book.active

    for row in range(13, sheet.max_row-2):
        if sheet[row][30].value != None:
            a = int(sheet[row][14].value)
            b = sheet[row][30].value
            sm += b
            if sheet[row][33].value != "Наличные":
                c = 2
            else:
                c = ""
            list_np.append([a, b, c])
    book.close()

    if os.path.exists("пул/bs2.xlsx"):
        book = openpyxl.open("пул/bs2.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs3.xlsx"):
        book = openpyxl.open("пул/bs3.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs4.xlsx"):
        book = openpyxl.open("пул/bs4.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs5.xlsx"):
        book = openpyxl.open("пул/bs5.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs6.xlsx"):
        book = openpyxl.open("пул/bs6.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs7.xlsx"):
        book = openpyxl.open("пул/bs7.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs8.xlsx"):
        book = openpyxl.open("пул/bs8.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    if os.path.exists("пул/bs9.xlsx"):
        book = openpyxl.open("пул/bs9.xlsx")
        sheet = book.active

        for row in range(13, sheet.max_row - 2):
            if sheet[row][30].value != None:
                a = int(sheet[row][14].value)
                b = sheet[row][30].value
                sm += b
                if sheet[row][33].value != "Наличные":
                    c = 2
                else:
                    c = ""
                list_np.append([a, b, c])
        book.close()

    #запись
    for i in range(1, len(list_np) + 1):
        # i принимает значения от 1 до 17, пишем со второй строки I+1, начиная с 0 индекса списка, 0 индекса подсписка I-1
        # 1,len(lst)+1 позваляет перебирать значения от 1 а не от 1, и включительно по число длины списка

        sheet_new[i + 1][0].value = list_np[i - 1][0]
        sheet_new[i + 1][7].value = list_np[i - 1][1]
        sheet_new[i + 1][8].value = list_np[i - 1][2]

    today = datetime.date.today()
    now = datetime.datetime.now()
    weekday = now.weekday()
    if weekday != 0:
        yesterday = today - datetime.timedelta(days=1)
    else:
        yesterday = today - datetime.timedelta(days=3)

    book_new.save("res/Отчет НП Боксберри РУ СЭЙФ от %s %s.xlsx" % (yesterday, round(sm, 2)))
    book_new.close()
